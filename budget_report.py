import csv
import os
from datetime import datetime
import openpyxl
from openpyxl.chart import BarChart,ProjectedPieChart,Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.worksheet.table import Table,TableStyleInfo
from openpyxl.worksheet.dimensions import ColumnDimension, DimensionHolder
from openpyxl.utils import get_column_letter
import re
import argparse

argParser = argparse.ArgumentParser()
argParser.add_argument("-b", "--budget", action="store_true", help="review and modify budget")
argParser.add_argument("-c", "--categorize", action="store_true", help="categorize unmapped transactions")
args = argParser.parse_args()

if args.categorize:
    bool_categorize_unmapped = True
else:
    bool_categorize_unmapped = False

dir = "/home/dev_iant/workspace/github.com/kaydencypress/budget/"
import_dir = dir + "import/"
category_map_file = dir + "category_map.csv"
categories_file = dir + "categories.txt"
outfile = dir + "export/export.xls"

class Transaction:
    def __init__(self,date,description,amount,type,category):
        self.date = date
        self.description = description
        self.amount = amount
        self.type = type
        self.category = category

class Menu:
    def __init__(self,prompt,menu_options):
        self.prompt = prompt
        self.menu_options = menu_options
        self.__prompt_user = False
        self.selection = None

    def print_menu(self):
        print("\n======================\n" + self.prompt)
        for option in self.menu_options:
            print(f"    {option[0]}: {option[1]}")
        return

    def get_user_input(self):
        self.__prompt_user = True
        while self.__prompt_user:
            self.print_menu()
            selection = input("Choose menu option: ")
            if selection.isalpha():
                selection = selection.upper()
            if selection not in dict(self.menu_options):
                print("Invalid input. Please enter the number/letter of one of the available menu options.")
            else:
                self.__prompt_user = False
        return selection

class Category:
    def __init__(self,name,budget):
        self.name = name
        self.budget = budget
        self.spending = []
    
    def calc_monthly_total(self,transactions):
        for transaction in transactions:
            if transaction.category == self.name:
                transaction_month = datetime.strptime(transaction.date,"%m/%d/%Y").strftime("%m-%Y")
                first_transacton = True
                for monthly_total in self.spending:
                    if monthly_total["month"] == transaction_month:
                        monthly_total["amount"] += transaction.amount
                        first_transacton = False
                        break
                if first_transacton:
                    self.spending.append({"month":transaction_month,"amount":transaction.amount})
        self.spending.sort(key=lambda x:x["amount"],reverse=True)

def import_transactions(dir):
    transactions = []
    try:
        for filename in os.listdir(dir):
            filepath = os.path.join(dir,filename)
            with open(filepath) as f:
                csv_reader = csv.reader(f)
                row_count = 0
                for row in csv_reader:
                    if filename.startswith("Chase"):
                        if row_count != 0:
                            date = row[0]
                            description = row[2]
                            type = row[4] 
                            amount = float(row[5])
                            tmp_category = row[3]
                            category = categorize_transaction(type,description,tmp_category)
                            transactions.append(Transaction(date,description,amount,type,category))
                    elif filename.startswith("Checking"):
                        date = row[0]
                        description = row[4]
                        amount = float(row[1])
                        type = "Checking"
                        category = categorize_transaction(type,description)
                        transactions.append(Transaction(date,description,amount,type,category))
                    elif filename.startswith("Savings"):
                        date = row[0]
                        description = row[4]
                        amount = float(row[1])
                        if "TRANSFER" in description.upper():
                            type = category = "Payment"
                        else:
                            transactions.append(Transaction(date,description,-amount,"Savings","Savings"))
                            if amount > 0:
                                transactions.append(Transaction(date,description,amount,"Income","Income"))
                    else:
                        print(f"Skipping invalid file: {filename}")
                    row_count+=1
    except Exception as e:
        print(e)
    return transactions

def read_category_csv(filepath):
    categories = []
    try:
        with open(filepath) as f:
            csv_reader = csv.reader(f)
            for row in csv_reader:
                category_budget = Category(row[0],float(row[1]))
                categories.append(category_budget)
    except Exception as e:
        print(e)
    return categories

def read_category_map(filepath):
    category_mapping = {}
    try:
        with open(filepath) as f:
            csv_reader = csv.reader(f)
            for row in csv_reader:
                category_mapping[row[0]] = row[1]
    except Exception as e:
        print(e)
    return category_mapping

def categorize_transaction(type,description,tmp_category=None):
    # map category from source data where possible
    if type == "Payment":
        return "Payment"
    elif type == "Income":
        return "Income"
    elif tmp_category == "Gas" or tmp_category == "Travel":
        return "Transportation"
    elif tmp_category == "Food & Drink" or tmp_category == "Groceries":
        return "Grocery"
    elif tmp_category == "Health & Wellness":
        return "Health"
    elif tmp_category == "Gifts & Donations":
        return "Donations"
    elif tmp_category == "Home":
        return "Other"

    # use custom mapping by description
    category_map = read_category_map(category_map_file)

    # parse transaction descriptions
    detail = None
    if "*" in description:
        description = description.split("*")
        detail = description[1]
        description = description[0]
    if "#" in description:
        description = description.split("#")[0]
    
    pattern = r'[0-9]|\/|\,'
    description = re.sub(pattern,'',description)
    description = description.rstrip()
    if "  " in description:
        description = description.split("  ")[0]

    # look up description for category match
    if detail:
        if detail in category_map:
            return category_map[detail]
    if description in category_map:
        return category_map[description]
    
    # prompt to create mapping if not present
    elif bool_categorize_unmapped:
        return categorize_unmapped_transactions(description,category_map_file,categories_file,detail)
    
    # default category if user has elected not to categorize unmapped transactions
    else:
        return "Other"

def map_description_or_detail(description,detail):
    if detail:
        # Prompt user whether to create categorize mapping based on description or detail
        menu_options = []
        menu_options.append(("1","Description"))
        menu_options.append(("2","Detail"))
        menu_options.append(("S","[Skip Expense]"))
        menu_options.append(("Q","[Quit Categorizing]"))
        prompt = f"Categorizing expense: {description} (Detail: {detail})\nCategorize based on description or detail?"
        description_detail_menu = Menu(prompt,menu_options)
        selection = description_detail_menu.get_user_input()
        # Evaluate user response
        if selection == "Q":
            global bool_categorize_unmapped
            bool_categorize_unmapped = False
            return
        if selection == "S":
            return
        if selection == "2":
            return detail
        if selection == "1":
            return description
    else:
        return description

def create_category_mapping(description,map_file,categories_file):
    categories = read_category_csv(categories_file)
    # Prompt user to map a category
    menu_options = []
    menu_num = 1
    for category in categories:
        menu_options.append((str(menu_num),category.name))
        menu_num+=1
    menu_options.append(("N","[Add New Category]"))
    menu_options.append(("S","[Skip Expense]"))
    menu_options.append(("X","[Ignore Expense from Reporting]"))
    menu_options.append(("Q","[Quit Categorizing]"))
    categorize_expense_menu = Menu("Create a category mapping for unmapped expense: " + description,menu_options)
    selection = categorize_expense_menu.get_user_input()
    # Evaluate user response
    # skip categorizing current transation and stop categorizing transactions
    if selection == "Q":
        global bool_categorize_unmapped
        bool_categorize_unmapped = False
        return "Other"
    # skip categorizing current transation
    elif selection == "S":
        return "Other"
    elif selection == "X":
        category = "Ignore"
    # prompt user and add new category name
    elif selection == "N":
        prompt_user = True
        while prompt_user:
            category = input("Enter name of new category: ")
            if category.isalnum():
                prompt_user = False
        with open(categories_file,'a') as f:
            f.write(category + ",0\n")
    # user selected an existing category
    else:
        category = categories[int(selection) - 1]
    # Add new mapping to category map file
    with open(map_file,'a') as f:
        row = description + "," + category.name + "\n"
        f.write(row)
    return category

def categorize_unmapped_transactions(description,map_file,categories_file,detail=None):
    str_to_evaluate = map_description_or_detail(description,detail)
    if str_to_evaluate:
        category = create_category_mapping(str_to_evaluate,map_file,categories_file)
    else:
        category = "Other"
    return category

def edit_budget(categories_file):
    # prompt user and add new category name
    prompt_user_main = True
    while prompt_user_main:
        categories = read_category_csv(categories_file)
        menu_options = []
        menu_num = 1
        for category in categories:
            menu_options.append((str(menu_num),f"{category.name} [${category.budget}]"))
            menu_num+=1
        menu_options.append(("N","[New Category]"))
        menu_options.append(("Q","[Quit Editing Budget]"))
        prompt = "Select category to edit monthly budget"
        budget_menu = Menu(prompt,menu_options)
        selection = budget_menu.get_user_input()
        if selection == "Q":
            prompt_user_main = False
            return
        elif selection == "N":
            category_name = input("Enter name of new category: ")
            if category_name.isalnum():
                selected_category = Category(category_name,0)
                categories.append(selected_category)
            else:
                print("Invalid entry: category name must be alphanumeric")
        else:
            selected_category = categories[int(selection)-1]

        prompt_user = True
        while prompt_user:
            new_budget = input(f"Enter new budget for {selected_category.name}: $")
            if new_budget == "":
                prompt_user = False
            try:
                new_budget = float(new_budget)
                prompt_user = False
            except:
                print("Invalid input, must be float")

        if new_budget:
            selected_category.budget = new_budget
            categories.sort(key=lambda x:x.budget,reverse=True)
            with open(categories_file,"w") as f:
                for category in categories:
                    f.write(f"{category.name},{category.budget}\n")
    return

def calc_overall_totals(categories,transactions):
    # TODO: totals = [
        #   {"month":"01-2023",
        #     "total_spending":1000,
        #      "total_income":3000
        # }]
    totals = []
    for category in categories:
        category.calc_monthly_total(transactions)
        for monthly_spend in category.spending:
            # add to sum of total income and expenses per month
            new_month = True
            if category.name == "Income":
                total_type = "total_income"
            else:
                total_type = "total_spending"
            for monthly_total in totals:
                if monthly_spend["month"] == monthly_total["month"]:
                    monthly_total[total_type] += monthly_spend["amount"]
                    new_month = False
            if new_month:
                if total_type == "total_income":
                    totals.append({"month":monthly_spend["month"],"total_income":monthly_spend["amount"],"total_spending":0})
                else:
                    totals.append({"month":monthly_spend["month"],"total_income":0,"total_spending":monthly_spend["amount"]})

    totals.sort(key=lambda x:datetime.strptime(x["month"],"%M-%Y"),reverse=False)
    return totals
        
def export_excel(categories,transactions,totals,filepath):
    wb = openpyxl.Workbook()
    wb.remove(wb["Sheet"])
    for category in categories:
        for monthly_spend in category.spending:
            # write data for monthly spending by category
            summary_sheet = create_sheet_if_needed(wb,monthly_spend["month"]+" Summary",["Category","Amount","Budget","Net"])
            row = [category.name,monthly_spend["amount"],category.budget,monthly_spend["amount"]+category.budget]
            if category.name != "Income":
                summary_sheet.append(row)    
            # write data for individual transaction details
            detail_sheet = create_sheet_if_needed(wb,monthly_spend["month"]+" Transactions",["Date","Amount","Description","Category","Type"])
            for transaction in transactions:
                if transaction.category != "Ignore" and  transaction.category != "Payment":
                    transaction_month = datetime.strptime(transaction.date,"%m/%d/%Y").strftime("%m-%Y")
                    if transaction_month == monthly_spend["month"]:
                        row = [transaction.date,transaction.amount,transaction.description,transaction.category,transaction.type]
                        detail_sheet.append(row)

    # write data for total income and expenses
    totals_sheet = create_sheet_if_needed(wb,"Monthly Totals",["Month","Income","Expenses","Net"])
    for monthly_total in totals:
        totals_sheet.append([monthly_total["month"],monthly_total["total_income"],-monthly_total["total_spending"],monthly_total["total_income"]+monthly_total["total_spending"]])
    totals_table = Table(displayName="total", ref=f"A1:D{totals_sheet.max_row}")
    totals_sheet.add_table(totals_table)

    # create bar chart for total income and expenses
    bar_chart = BarChart()
    bar_chart.width = 20
    bar_chart.height = 10
    data_range = Reference(totals_sheet,min_col=2,min_row=1,max_col=3,max_row=len(totals))
    label_range = Reference(totals_sheet,min_col=1,min_row=2,max_col=1,max_row=len(totals))
    bar_chart.add_data(data_range,titles_from_data=True)
    bar_chart.set_categories(label_range)
    bar_chart.title = "Income and Expenses"
    totals_sheet.add_chart(bar_chart,"F1")

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        # adjust column widths
        for column_cells in ws.columns:
            length = max(len(cell_value_str(cell.value)) for cell in column_cells) + 5
            ws.column_dimensions[column_cells[0].column_letter].width = length
        # create tables for data
        table_range = "A1:" + get_column_letter(ws.max_column) + str(ws.max_row)
        data_table = Table(displayName=sheet_name.replace(" ","_"), ref=table_range)
        data_table.tableStyleInfo = TableStyleInfo(name="TableStyleLight13",showRowStripes=True, showColumnStripes=False,showFirstColumn=False,showLastColumn=False)
        ws.add_table(data_table)
        if "SUMMARY" in sheet_name.upper():
            # create chart for monthly spending by category
            projected_pie = ProjectedPieChart()
            projected_pie.type = 'bar'
            projected_pie.width = 20
            projected_pie.height = 10
            data_range = Reference(ws,min_col=2,min_row=2,max_col=2,max_row=ws.max_row)
            label_range = Reference(ws,min_col=1,min_row=2,max_col=1,max_row=ws.max_row)
            projected_pie.add_data(data_range,titles_from_data=False)
            projected_pie.set_categories(label_range)
            projected_pie.dataLabels = DataLabelList()
            projected_pie.dataLabels.showPercent = True
            projected_pie.dataLabels.showLeaderLines = True
            projected_pie.dataLabels.showCatName = True
            projected_pie.dataLabels.separator = ','
            projected_pie.title = "Expenses by Category"
            projected_pie.legend.position = 'b'
            ws.add_chart(projected_pie,"F1")

    # save Excel sheet
    try:
        wb.save(filepath)
        print(f"Wrote output report: {filepath}")
    except Exception as e:
        print(f"Error: unable to write output report: {e}")
    return

def create_sheet_if_needed(wb,title,header):
    if title not in wb.sheetnames:
        try:
            new_sheet_date = datetime.strptime(title[0:7],"%m-%Y")
            i = 0
            for sheet_name in wb.sheetnames:
                current_sheet_date = datetime.strptime(sheet_name[0:7],"%m-%Y")
                # find index for where to put new sheet
                if new_sheet_date > current_sheet_date:
                    break
                elif new_sheet_date < current_sheet_date:
                    i+=1
                    continue
                elif "SUMMARY" in title.upper():
                    break
                else:
                    i+=1
                    break
            ws = wb.create_sheet(title,index=i)
        except:
            # add sheet to the beginning if not for a specific month
            ws = wb.create_sheet(title,index=0)
        
        ws.append(header)
    else:
        ws = wb[title]
    return ws

def cell_value_str(value):
    if value is None:
        return ""
    if isinstance(value, float):
        return "%.2f" % value
    return str(value)

def main():
    if args.budget:
        edit_budget(categories_file)
    categories = read_category_csv(categories_file)
    transactions = import_transactions(import_dir)
    totals = calc_overall_totals(categories,transactions)
    export_excel(categories,transactions,totals,outfile)

main()